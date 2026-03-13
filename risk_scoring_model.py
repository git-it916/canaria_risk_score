"""
Global Multi-Asset Risk Scoring Model
======================================
Measures the daily market Risk On / Risk Off regime using a weighted
composite score derived from "Canary" asset classes fetched live from
Bloomberg Terminal via the xbbg library.

Usage
-----
    from risk_scoring_model import RiskScoringModel

    model = RiskScoringModel()
    results = model.run(start_date="2015-01-01", end_date="2024-12-31")

    # Step-by-step (same result)
    model.fetch_data("2015-01-01", "2024-12-31")
    model.calculate_signals()
    results = model.get_regime()   # pd.DataFrame ["score", "regime"]
    model.plot_results()

    # Inspect individual signals
    print(model.signal_summary())
"""

from __future__ import annotations

import os
import warnings
from datetime import datetime, timedelta

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.gridspec import GridSpec
from matplotlib.lines import Line2D
from matplotlib.patches import Patch

warnings.filterwarnings("ignore")

# ── Bloomberg availability guard ──────────────────────────────────────────────
try:
    from xbbg import blp

    _BBG_AVAILABLE = True
except ImportError:
    _BBG_AVAILABLE = False
    print(
        "[WARN] xbbg / Bloomberg Terminal not available.\n"
        "       Install with: pip install xbbg\n"
        "       The model will run in DEMO mode with synthetic data."
    )


# ── Synthetic data helper (offline / CI testing) ──────────────────────────────

def _make_synthetic_prices(
    tickers: list[str],
    start: str,
    end: str,
    base: float = 100.0,
    vol: float = 0.01,
    seed: int = 0,
) -> pd.DataFrame:
    """Return a deterministic random-walk price DataFrame for offline use."""
    rng = np.random.default_rng(seed)
    idx = pd.bdate_range(start, end)
    out: dict[str, np.ndarray] = {}
    for i, t in enumerate(tickers):
        r = rng.normal(0.0, vol, size=len(idx))
        out[t] = (base + i * 5) * np.exp(np.cumsum(r))
    return pd.DataFrame(out, index=idx)


# ══════════════════════════════════════════════════════════════════════════════

class RiskScoringModel:
    """
    Global Multi-Asset Risk Scoring Model.

    The model aggregates weighted boolean signals from twelve market
    "canaries" plus three macro filters into a single daily composite
    score, then classifies each day as Risk On / Transition / Risk Off
    using percentile-based thresholds on the full historical sample.

    Attributes set after running the pipeline
    -----------------------------------------
    signals  : pd.DataFrame  — daily weighted score contribution per asset
    scores   : pd.Series     — daily composite risk score (sum of signals)
    regimes  : pd.Series     — daily regime label string
    spx_data : pd.DataFrame  — S&P 500 prices used for the overlay chart
    """

    # ── Asset universe ────────────────────────────────────────────────────────
    #
    # PRICE_ASSETS  — PX_LAST signals driven by a rolling MA crossover.
    #   Format: ticker → (ma_period, weight, direction, label)
    #   direction "above" : Price > MA  →  Risk On (+weight)
    #   direction "below" : Price < MA  →  Risk On (+weight)
    #
    PRICE_ASSETS: dict[str, tuple] = {
        "DXY Curncy":    (10, 2.0, "below", "Dollar Index"),      # weak USD → risk-on
        "EMB US Equity": (10, 1.0, "above", "EM Bond"),
        "CEW US Equity": (10, 1.0, "above", "EM Currency"),
        "BND US Equity": (10, 1.0, "below", "US Treasury Agg"),   # rates rising → risk-on context
        "TIP US Equity": (10, 1.0, "below", "TIPS"),
        "VEA US Equity": (10, 1.0, "above", "Dev. Market Eq"),
        "GLD US Equity": (10, 1.0, "below", "Gold"),              # safe-haven outflow → risk-on
        "DBC US Equity": (10, 1.0, "above", "Commodity"),
        "VIX Index":     (60, 2.0, "below", "VIX"),               # high sensitivity
        "VVIX Index":    (60, 1.0, "below", "VVIX"),
    }

    # SKEW: pure threshold — no MA.  When SKEW > threshold the score is
    # penalised by SKEW_PENALTY (negative number).
    SKEW_TICKER:    str   = "SKEW Index"
    SKEW_THRESHOLD: float = 140.0
    SKEW_PENALTY:   float = -2.0   # e.g. –2 points when tail risk is elevated

    # SPREAD_ASSETS — OAS-based credit signals.
    #   Format: ticker → (bbg_field, ma_period, weight, direction, label)
    #   Fallback fields are tried in order if primary field is unavailable.
    SPREAD_ASSETS: dict[str, tuple] = {
        "LF98TRUU Index": ("OAS_SPREAD_BID", 60, 2.0, "below", "US HY OAS"),
    }
    # Alternative OAS field names tried in order when primary fails
    _OAS_FALLBACK_FIELDS: list[str] = [
        "OAS_SPREAD_BID", "OAS_BID", "OAS_MID", "OAS_SPREAD",
        "OAS", "Z_SPREAD_MID", "OPTION_ADJ_SPREAD",
    ]

    # MACRO_ASSETS — monthly/quarterly releases; forward-filled to daily.
    #   Format: ticker → (condition, threshold, weight, label)
    #   condition "above" : value > threshold  →  Risk On (+weight)
    #   condition "below" : value < threshold  →  Risk On (+weight)
    MACRO_ASSETS: dict[str, tuple] = {
        "NAPMPMI Index": ("above", 50.0, 1.0, "ISM PMI"),
        "CPI YOY Index": ("below",  3.0, 1.0, "CPI YoY"),
        "USURTOT Index": ("below",  4.0, 1.0, "Unemployment"),
    }

    SPX_TICKER: str = "SPX Index"

    # ── Regime classification ─────────────────────────────────────────────────
    RISK_ON_PCT:  int = 70   # score ≥ p70  →  Risk On
    RISK_OFF_PCT: int = 30   # score ≤ p30  →  Risk Off

    # Calendar days of extra history fetched before start_date for MA warm-up.
    # 60-day MA + buffer = 120 days is ample.
    MA_WARMUP_DAYS: int = 120

    # ── Colour palette ────────────────────────────────────────────────────────
    PALETTE: dict[str, str] = {
        "Risk On":    "#27ae60",
        "Transition": "#e67e22",
        "Risk Off":   "#c0392b",
        "score":      "#2c3e50",
        "score_ma":   "#8e44ad",
        "spx":        "#2980b9",
    }

    # ═════════════════════════════════════════════════════════════════════════
    # Constructor
    # ═════════════════════════════════════════════════════════════════════════

    def __init__(self) -> None:
        # Internal stores (contain extended warm-up history)
        self._full_price:  pd.DataFrame | None = None
        self._full_spread: pd.DataFrame | None = None
        self._full_macro:  pd.DataFrame | None = None

        self._start_date: str | None = None
        self._end_date:   str | None = None

        # Public outputs (trimmed to [start_date, end_date])
        self.signals:  pd.DataFrame | None = None
        self.scores:   pd.Series     | None = None
        self.regimes:  pd.Series     | None = None
        self.spx_data: pd.DataFrame  | None = None

    # ═════════════════════════════════════════════════════════════════════════
    # 1. fetch_data
    # ═════════════════════════════════════════════════════════════════════════

    def fetch_data(self, start_date: str, end_date: str) -> None:
        """
        Pull all required data from Bloomberg Terminal via xbbg.

        An extra MA_WARMUP_DAYS buffer is prepended to *start_date* so that
        rolling averages (up to 60-day) are fully initialised on the first
        date of the requested window.  The extended data is stored in private
        attributes; ``calculate_signals`` trims the final signals to
        [start_date, end_date].

        Parameters
        ----------
        start_date : str   "YYYY-MM-DD"
        end_date   : str   "YYYY-MM-DD"
        """
        self._start_date = start_date
        self._end_date   = end_date

        # Extended start for warm-up
        ext_start = (
            datetime.strptime(start_date, "%Y-%m-%d")
            - timedelta(days=self.MA_WARMUP_DAYS)
        ).strftime("%Y-%m-%d")

        _sep = "─" * 62
        print(f"\n{_sep}")
        print("  BLOOMBERG DATA FETCH")
        print(_sep)
        print(f"  Requested window : {start_date}  →  {end_date}")
        print(f"  Fetch window     : {ext_start}  →  {end_date}"
              f"  (+{self.MA_WARMUP_DAYS}d MA warm-up)")

        if _BBG_AVAILABLE:
            self._full_price  = self._fetch_prices(ext_start, end_date)
            self._full_spread = self._fetch_spreads(ext_start, end_date)
            self._full_macro  = self._fetch_macro(ext_start, end_date)
            self.spx_data     = self._fetch_spx(start_date, end_date)
        else:
            print("\n  [DEMO] Generating synthetic data …")
            self._load_synthetic_data(ext_start, end_date, start_date)

        print(f"{_sep}\n")

    # ── Private Bloomberg fetch helpers ───────────────────────────────────────

    def _fetch_prices(self, ext_start: str, end_date: str) -> pd.DataFrame:
        """Fetch PX_LAST for all price-based assets and SKEW."""
        tickers = list(self.PRICE_ASSETS.keys()) + [self.SKEW_TICKER]
        print(f"\n  [1/4] Price assets  ({len(tickers)} tickers) …")

        df = blp.bdh(
            tickers=tickers,
            flds="PX_LAST",
            start_date=ext_start,
            end_date=end_date,
        )
        df = self._flatten_bbg_columns(df)
        df = df.ffill()

        # Report any tickers the terminal returned no data for
        missing = [t for t in tickers if t not in df.columns]
        if missing:
            print(f"       [WARN] No data returned for: {missing}")
        print(f"       → {df.shape[0]} rows × {df.shape[1]} cols")
        return df

    def _fetch_spreads(self, ext_start: str, end_date: str) -> pd.DataFrame:
        """
        Fetch OAS / spread fields for credit assets.

        Bloomberg field names for the HY OAS can vary by terminal licence.
        We try OAS_SPREAD_BID → OAS_BID → OAS_SPREAD → OAS in order and
        use the first one that returns non-empty data.
        """
        print(f"\n  [2/4] Spread assets ({len(self.SPREAD_ASSETS)} tickers) …")
        frames: list[pd.DataFrame] = []

        for ticker, (primary_field, *_) in self.SPREAD_ASSETS.items():
            fields_to_try = [primary_field] + [
                f for f in self._OAS_FALLBACK_FIELDS if f != primary_field
            ]
            fetched = False
            for fld in fields_to_try:
                try:
                    raw = blp.bdh(
                        tickers=ticker,
                        flds=fld,
                        start_date=ext_start,
                        end_date=end_date,
                    )
                    if raw is not None and not raw.empty:
                        raw = self._flatten_bbg_columns(raw)
                        raw.columns = [ticker]
                        raw = raw.ffill()
                        frames.append(raw)
                        print(f"       {ticker} → fetched with field '{fld}'"
                              f"  ({raw.shape[0]} rows)")
                        fetched = True
                        break
                except Exception as exc:  # noqa: BLE001
                    print(f"         '{fld}' → {exc}")  # show reason, try next

            if not fetched:
                print(f"       [WARN] Could not fetch {ticker} — "
                      f"tried fields {fields_to_try}.  Spread signal skipped.")

        if frames:
            return pd.concat(frames, axis=1)
        return pd.DataFrame()

    def _fetch_macro(self, ext_start: str, end_date: str) -> pd.DataFrame:
        """
        Fetch monthly macro indicators.

        BDH returns values only on release/publication dates (sparse).
        The sparse frame is stored here and forward-filled to daily in
        ``calculate_signals``.
        """
        tickers = list(self.MACRO_ASSETS.keys())
        print(f"\n  [3/4] Macro assets  ({len(tickers)} tickers) …")
        try:
            df = blp.bdh(
                tickers=tickers,
                flds="PX_LAST",
                start_date=ext_start,
                end_date=end_date,
            )
            df = self._flatten_bbg_columns(df)
            # Drop rows that are all-NaN (typical for monthly series on daily BDH)
            df = df.dropna(how="all")
            print(f"       → {df.shape[0]} observations (sparse monthly)")
            return df
        except Exception as exc:
            print(f"       [WARN] Macro fetch failed: {exc}")
            return pd.DataFrame()

    def _fetch_spx(self, start_date: str, end_date: str) -> pd.DataFrame:
        """Fetch S&P 500 for the overlay chart (no warm-up needed)."""
        print(f"\n  [4/4] Benchmark     ({self.SPX_TICKER}) …")
        try:
            df = blp.bdh(
                tickers=self.SPX_TICKER,
                flds="PX_LAST",
                start_date=start_date,
                end_date=end_date,
            )
            df = self._flatten_bbg_columns(df)
            df.columns = [self.SPX_TICKER]
            df = df.ffill()
            print(f"       → {df.shape[0]} rows")
            return df
        except Exception as exc:
            print(f"       [WARN] SPX fetch failed: {exc}")
            return pd.DataFrame()

    # ── Offline demo mode ─────────────────────────────────────────────────────

    def _load_synthetic_data(
        self, ext_start: str, end_date: str, start_date: str
    ) -> None:
        """Populate all data stores with deterministic synthetic data."""
        # ── Price assets ──
        price_tickers = list(self.PRICE_ASSETS.keys()) + [self.SKEW_TICKER]
        self._full_price = _make_synthetic_prices(
            price_tickers, ext_start, end_date, seed=1
        )
        n = len(self._full_price)
        t = np.linspace(0, 6 * np.pi, n)

        # Override specific tickers with more realistic dynamics
        # VIX: mean-reverting, range 10–40
        self._full_price["VIX Index"]  = 18 + 10 * np.abs(np.sin(t * 0.7))
        # VVIX: range 70–120
        self._full_price["VVIX Index"] = 90 + 18 * np.abs(np.sin(t * 0.5 + 1))
        # SKEW: oscillates around threshold (130–155)
        self._full_price[self.SKEW_TICKER] = 140 + 12 * np.sin(t * 0.4)
        # DXY: range 90–110
        self._full_price["DXY Curncy"] = 100 + 8 * np.sin(t * 0.3 + 0.5)

        # ── Spread assets ──
        spread_tickers = list(self.SPREAD_ASSETS.keys())
        self._full_spread = _make_synthetic_prices(
            spread_tickers, ext_start, end_date, base=350, vol=0.008, seed=2
        )
        # HY OAS: range 250–600 bps
        self._full_spread["LF98TRUU Index"] = (
            400 + 150 * np.sin(t[: len(self._full_spread)] * 0.35)
        )

        # ── Macro assets (monthly cadence) ──
        monthly_idx = pd.date_range(ext_start, end_date, freq="MS")
        m = len(monthly_idx)
        tm = np.linspace(0, 4 * np.pi, m)
        self._full_macro = pd.DataFrame(
            {
                "NAPMPMI Index": 50.5 + 4.0 * np.sin(tm),
                "CPI YOY Index": 2.8  + 1.5 * np.sin(tm + 1.0),
                "USURTOT Index": 3.9  + 0.6 * np.sin(tm + 2.0),
            },
            index=monthly_idx,
        )
        # Convert to business-day index (drop weekends — Bloomberg does the same)
        bday_mask = self._full_macro.index.day_of_week < 5
        self._full_macro = self._full_macro[bday_mask]

        # ── SPX ──
        self.spx_data = _make_synthetic_prices(
            [self.SPX_TICKER], start_date, end_date, base=2000, vol=0.009, seed=3
        )

    # ── Utility ───────────────────────────────────────────────────────────────

    @staticmethod
    def _flatten_bbg_columns(df: pd.DataFrame) -> pd.DataFrame:
        """
        Collapse xbbg's MultiIndex columns ``(ticker, field)`` → ``ticker``.

        xbbg returns MultiIndex columns when multiple tickers or fields are
        requested.  We always want the ticker (level 0) as the column label.
        """
        if isinstance(df.columns, pd.MultiIndex):
            df = df.copy()
            df.columns = df.columns.get_level_values(0)
        # xbbg sometimes returns datetime.date objects instead of pd.Timestamp.
        # Convert to DatetimeIndex so .loc["YYYY-MM-DD"] slicing works correctly.
        df.index = pd.to_datetime(df.index)
        return df

    # ═════════════════════════════════════════════════════════════════════════
    # 2. calculate_signals
    # ═════════════════════════════════════════════════════════════════════════

    def calculate_signals(self) -> pd.DataFrame:
        """
        Compute rolling MAs and derive weighted daily signals for every asset.

        MA computation is performed on the full extended-history data so that
        values on the first requested date are already warmed up.  The final
        ``self.signals`` DataFrame is trimmed to [start_date, end_date].

        Signal convention
        -----------------
        * Regular asset  : +weight  when risk-on condition holds, else 0
        * SKEW penalty   : SKEW_PENALTY (< 0) when SKEW > 140, else 0

        Returns
        -------
        pd.DataFrame  rows = business dates in [start, end],
                      columns = human-readable signal labels
        """
        if self._full_price is None:
            raise RuntimeError("No data loaded — call fetch_data() first.")

        frames: list[pd.Series] = []

        # ── 2a. Price-based MA signals ────────────────────────────────────
        for ticker, (ma_period, weight, direction, label) in self.PRICE_ASSETS.items():
            series: pd.Series | None = self._full_price.get(ticker)
            if series is None or series.dropna().empty:
                print(f"  [WARN] {ticker} unavailable — signal skipped.")
                continue

            ma = series.rolling(window=ma_period, min_periods=ma_period).mean()
            cond = (series > ma) if direction == "above" else (series < ma)
            weighted = cond.astype(float) * weight
            weighted.name = label
            frames.append(weighted)

        # ── 2b. SKEW tail-risk penalty ────────────────────────────────────
        skew_s: pd.Series | None = self._full_price.get(self.SKEW_TICKER)
        if skew_s is not None and not skew_s.dropna().empty:
            # Contribution = SKEW_PENALTY (<0) when SKEW > threshold, else 0
            penalty_vals = np.where(
                skew_s > self.SKEW_THRESHOLD, self.SKEW_PENALTY, 0.0
            )
            penalty = pd.Series(
                penalty_vals, index=skew_s.index, name="SKEW Penalty"
            )
            frames.append(penalty)
        else:
            print(f"  [WARN] {self.SKEW_TICKER} unavailable — penalty skipped.")

        # ── 2c. Spread-based MA signals ───────────────────────────────────
        if self._full_spread is not None and not self._full_spread.empty:
            for ticker, (_, ma_period, weight, direction, label) in \
                    self.SPREAD_ASSETS.items():
                series = self._full_spread.get(ticker)
                if series is None or series.dropna().empty:
                    print(f"  [WARN] {ticker} spread unavailable — skipped.")
                    continue
                ma = series.rolling(window=ma_period, min_periods=ma_period).mean()
                cond = (series > ma) if direction == "above" else (series < ma)
                weighted = cond.astype(float) * weight
                weighted.name = label
                frames.append(weighted)

        # ── 2d. Macro signals (monthly → daily via ffill) ─────────────────
        if self._full_macro is not None and not self._full_macro.empty:
            # Target daily timeline: business days within the extended window
            daily_idx = self._full_price.index

            for ticker, (cond_dir, threshold, weight, label) in \
                    self.MACRO_ASSETS.items():
                series = self._full_macro.get(ticker)
                if series is None or series.dropna().empty:
                    print(f"  [WARN] {ticker} macro data missing — skipped.")
                    continue

                # Reindex to daily then forward-fill release values
                daily_s = (
                    series
                    .dropna()
                    .reindex(daily_idx.union(series.dropna().index))
                    .ffill()
                    .reindex(daily_idx)
                )
                cond = (
                    (daily_s > threshold) if cond_dir == "above"
                    else (daily_s < threshold)
                )
                weighted = cond.astype(float) * weight
                weighted.name = label
                frames.append(weighted)

        if not frames:
            raise RuntimeError(
                "No signals were computed — check data availability."
            )

        # Combine, sort and trim to user-requested window
        all_signals = pd.concat(frames, axis=1).sort_index()
        self.signals = (
            all_signals
            .loc[self._start_date : self._end_date]
            .copy()
        )

        # Forward-fill only: macro signals may have NaN for the very first rows
        # if the warmup window didn't include a prior release date.
        # bfill() is intentionally omitted to avoid look-ahead bias.
        self.signals = self.signals.ffill()

        return self.signals

    # ═════════════════════════════════════════════════════════════════════════
    # 3. get_regime
    # ═════════════════════════════════════════════════════════════════════════

    def get_regime(self) -> pd.DataFrame:
        """
        Aggregate signals into a composite score and classify each day.

        Regime thresholds are derived from the *in-sample* percentile
        distribution so the labelling adapts automatically to any date range:

        ┌──────────────┬────────────────────────────────────────────┐
        │  Risk On     │ score ≥ p70 of full-sample distribution     │
        │  Transition  │ p30 < score < p70                           │
        │  Risk Off    │ score ≤ p30                                  │
        └──────────────┴────────────────────────────────────────────┘

        Returns
        -------
        pd.DataFrame  with columns ["score", "regime"]
        """
        if self.signals is None:
            self.calculate_signals()

        # Composite score = sum of all weighted signal contributions
        self.scores = self.signals.sum(axis=1, min_count=1)
        self.scores.name = "Total Risk Score"

        p_on  = float(self.scores.quantile(self.RISK_ON_PCT  / 100))
        p_off = float(self.scores.quantile(self.RISK_OFF_PCT / 100))

        def _classify(s: float) -> str:
            if pd.isna(s):
                return "Unknown"
            if s >= p_on:
                return "Risk On"
            if s <= p_off:
                return "Risk Off"
            return "Transition"

        self.regimes = self.scores.apply(_classify)
        self.regimes.name = "Regime"

        result = pd.concat([self.scores, self.regimes], axis=1)
        result.columns = ["score", "regime"]

        self._print_regime_summary(p_on, p_off)
        return result

    def _print_regime_summary(self, p_on: float, p_off: float) -> None:
        """Console summary printed after get_regime()."""
        sep = "═" * 62
        print(f"\n{sep}")
        print("  GLOBAL RISK SCORING MODEL — REGIME SUMMARY")
        print(sep)
        print(f"  Period            : {self._start_date}  →  {self._end_date}")
        print(f"  Score range       : {self.scores.min():.2f}  →  "
              f"{self.scores.max():.2f}")
        print(f"  Mean  / Std       : {self.scores.mean():.2f}  /  "
              f"{self.scores.std():.2f}")
        print(f"  Risk On  (≥p{self.RISK_ON_PCT})  : {p_on:.2f}")
        print(f"  Risk Off (≤p{self.RISK_OFF_PCT})  : {p_off:.2f}")
        print()

        counts = self.regimes.value_counts()
        total  = counts.sum()
        order  = ["Risk On", "Transition", "Risk Off", "Unknown"]
        for regime in order:
            if regime not in counts:
                continue
            cnt = counts[regime]
            pct = cnt / total * 100
            bar = "█" * int(pct / 2)          # 50 chars = 100 %
            col = self.PALETTE.get(regime, "grey")
            print(f"  {regime:<12} {cnt:5d} days  {pct:5.1f}%  {bar}")

        latest_date  = self.scores.index[-1].date()
        latest_score = self.scores.iloc[-1]
        latest_reg   = self.regimes.iloc[-1]
        print(f"\n  Latest ({latest_date}) : "
              f"Score = {latest_score:.2f}   →   [ {latest_reg} ]")
        print(f"{sep}\n")

    # ═════════════════════════════════════════════════════════════════════════
    # 4. plot_results
    # ═════════════════════════════════════════════════════════════════════════

    def plot_results(
        self,
        figsize: tuple[int, int] = (18, 11),
        save_path: str | None = "risk_score_chart.png",
    ) -> None:
        """
        Three-panel figure:

        Panel 1 — S&P 500 price with regime-coloured background shading.
        Panel 2 — Composite Risk Score with 20-day smoothed MA and
                  percentile threshold lines.
        Panel 3 — Stacked area of individual weighted signal contributions
                  (positive signals stacked upward, SKEW penalty shaded red).

        Parameters
        ----------
        figsize   : (width, height) in inches
        save_path : file path to save PNG; ``None`` to skip saving
        """
        if self.scores is None or self.regimes is None:
            raise RuntimeError("Call get_regime() before plot_results().")

        idx   = self.scores.index
        C     = self.PALETTE
        p_on  = float(self.scores.quantile(self.RISK_ON_PCT  / 100))
        p_off = float(self.scores.quantile(self.RISK_OFF_PCT / 100))

        # ── Figure layout ─────────────────────────────────────────────────
        fig = plt.figure(figsize=figsize, facecolor="#f7f8fa")
        fig.suptitle(
            "Global Multi-Asset Risk Scoring Model",
            fontsize=15, fontweight="bold", y=0.995,
        )
        gs = GridSpec(
            3, 1, figure=fig,
            height_ratios=[2.4, 2.0, 1.2],
            hspace=0.05,
        )
        ax1 = fig.add_subplot(gs[0])
        ax2 = fig.add_subplot(gs[1], sharex=ax1)
        ax3 = fig.add_subplot(gs[2], sharex=ax1)

        # ── Shared regime shading ──────────────────────────────────────────
        for ax in (ax1, ax2, ax3):
            self._shade_regimes(ax, idx, self.regimes, C)

        # ─────────────────────────────────────────────────────────────────
        # Panel 1: S&P 500
        # ─────────────────────────────────────────────────────────────────
        if self.spx_data is not None and not self.spx_data.empty:
            spx_aligned = self.spx_data.iloc[:, 0].reindex(idx).ffill()
            ax1.plot(idx, spx_aligned, color=C["spx"], lw=1.8, label="S&P 500 (SPX)")

        ax1.set_ylabel("SPX Level", fontsize=10)
        ax1.grid(True, alpha=0.25, linestyle="--")
        ax1.set_title("S&P 500  ·  Risk Regime Background", fontsize=10, pad=4)

        # Regime legend patches
        regime_patches = [
            Patch(facecolor=C["Risk On"],    alpha=0.30, label="Risk On"),
            Patch(facecolor=C["Transition"], alpha=0.30, label="Transition"),
            Patch(facecolor=C["Risk Off"],   alpha=0.30, label="Risk Off"),
        ]
        spx_line = Line2D([0], [0], color=C["spx"], lw=1.8, label="S&P 500")
        ax1.legend(
            handles=[spx_line] + regime_patches,
            loc="upper left", fontsize=8, framealpha=0.7,
        )
        plt.setp(ax1.get_xticklabels(), visible=False)

        # ─────────────────────────────────────────────────────────────────
        # Panel 2: Composite Risk Score
        # ─────────────────────────────────────────────────────────────────
        ax2.plot(idx, self.scores, color=C["score"], lw=1.6,
                 label="Risk Score", zorder=4)

        ma20 = self.scores.rolling(20, min_periods=5).mean()
        ax2.plot(idx, ma20, color=C["score_ma"], lw=1.2, ls="--",
                 alpha=0.85, label="20-day MA", zorder=4)

        ax2.axhline(p_on,  color=C["Risk On"],  ls=":", lw=1.5, alpha=0.9,
                    label=f"Risk On  threshold  (p{self.RISK_ON_PCT} = {p_on:.1f})")
        ax2.axhline(p_off, color=C["Risk Off"], ls=":", lw=1.5, alpha=0.9,
                    label=f"Risk Off threshold  (p{self.RISK_OFF_PCT} = {p_off:.1f})")
        ax2.axhline(0, color="black", lw=0.5, alpha=0.25)

        # Fill between score and p_on/p_off to highlight excursions
        ax2.fill_between(idx, self.scores, p_on,
                          where=(self.scores >= p_on),
                          alpha=0.08, color=C["Risk On"], zorder=2)
        ax2.fill_between(idx, self.scores, p_off,
                          where=(self.scores <= p_off),
                          alpha=0.08, color=C["Risk Off"], zorder=2)

        ax2.set_ylabel("Composite Score", fontsize=10)
        ax2.grid(True, alpha=0.25, linestyle="--")
        ax2.legend(loc="upper left", fontsize=8, ncol=2, framealpha=0.7)
        plt.setp(ax2.get_xticklabels(), visible=False)

        # ─────────────────────────────────────────────────────────────────
        # Panel 3: Signal contribution breakdown (stacked area)
        # ─────────────────────────────────────────────────────────────────
        pos_cols = [c for c in self.signals.columns if "Penalty" not in c]
        neg_cols = [c for c in self.signals.columns if "Penalty"     in c]

        cmap = plt.cm.tab20
        pos_arrays = [
            self.signals[c].reindex(idx).fillna(0).values
            for c in pos_cols
        ]
        colors_pos = [cmap(i / max(len(pos_cols), 1)) for i in range(len(pos_cols))]

        if pos_arrays:
            ax3.stackplot(
                idx, pos_arrays,
                labels=pos_cols,
                colors=colors_pos,
                alpha=0.72,
            )

        # Penalty (negative contribution) — red fill downward from 0
        for col in neg_cols:
            vals = self.signals[col].reindex(idx).fillna(0)
            ax3.fill_between(
                idx, vals, 0,
                color=C["Risk Off"], alpha=0.65, label=col,
            )

        ax3.axhline(0, color="black", lw=0.7, alpha=0.5)
        ax3.set_ylabel("Signal Contrib.", fontsize=9)
        ax3.grid(True, alpha=0.20, linestyle="--")

        # ── X-axis tick format ─────────────────────────────────────────────
        n_days = len(idx)
        interval = max(1, round(n_days / 252 / 4))   # ~4 ticks per year
        ax3.xaxis.set_major_locator(mdates.MonthLocator(interval=interval * 3))
        ax3.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
        plt.setp(ax3.get_xticklabels(), rotation=35, ha="right", fontsize=8)

        # ── Save & show ────────────────────────────────────────────────────
        plt.tight_layout()
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches="tight")
            print(f"[INFO] Chart saved → {save_path}")
        plt.show()

    # ── Regime shading helper ─────────────────────────────────────────────────

    @staticmethod
    def _shade_regimes(
        ax: plt.Axes,
        idx: pd.DatetimeIndex,
        regimes: pd.Series,
        colors: dict[str, str],
    ) -> None:
        """Fill axis background with translucent regime-coloured bands."""
        if len(idx) < 2:
            return

        # Walk through regime transitions, collecting (start, end, regime) spans
        segments: list[tuple] = []
        cur_regime = regimes.iloc[0]
        seg_start  = idx[0]

        for date, regime in zip(idx[1:], regimes.iloc[1:]):
            if regime != cur_regime:
                segments.append((seg_start, date, cur_regime))
                cur_regime = regime
                seg_start  = date
        segments.append((seg_start, idx[-1], cur_regime))

        for s, e, r in segments:
            ax.axvspan(
                s, e,
                alpha=0.12,
                color=colors.get(r, "#cccccc"),
                linewidth=0,
            )

    # ═════════════════════════════════════════════════════════════════════════
    # 5. Convenience: end-to-end pipeline
    # ═════════════════════════════════════════════════════════════════════════

    def run(
        self,
        start_date: str | None = None,
        end_date:   str | None = None,
        plot:       bool = True,
    ) -> pd.DataFrame:
        """
        Execute the full pipeline in one call:
        fetch_data → calculate_signals → get_regime → [plot_results].

        Parameters
        ----------
        start_date : "YYYY-MM-DD" or None  → 기본값: 오늘로부터 2년 전
        end_date   : "YYYY-MM-DD" or None  → 기본값: 오늘
        plot       : bool  whether to display the three-panel chart

        Returns
        -------
        pd.DataFrame  with columns ["score", "regime"]
        """
        today = pd.Timestamp.now().normalize()
        if end_date is None:
            end_date = today.strftime("%Y-%m-%d")
        if start_date is None:
            start_date = (today - pd.DateOffset(years=2)).strftime("%Y-%m-%d")

        self.fetch_data(start_date, end_date)
        self.calculate_signals()
        results = self.get_regime()
        if plot:
            self.plot_results()
        return results

    # ── Diagnostic helper ─────────────────────────────────────────────────────

    def signal_summary(self, date: str | None = None) -> pd.DataFrame:
        """
        Return a table of every signal's value and status on a given date.

        Parameters
        ----------
        date : "YYYY-MM-DD" or None  → uses the last available date

        Returns
        -------
        pd.DataFrame  indexed by signal name with columns
                      ["Weight", "Value", "Status"]
        """
        if self.signals is None:
            raise RuntimeError("Call calculate_signals() first.")

        if date is None:
            row = self.signals.iloc[-1]
            dt  = self.signals.index[-1].date()
        else:
            row = self.signals.loc[date]
            dt  = date

        # Build weight lookup from configs
        weight_map: dict[str, float] = {}
        for _, (_, w, _, lbl) in self.PRICE_ASSETS.items():
            weight_map[lbl] = w
        weight_map["SKEW Penalty"] = self.SKEW_PENALTY
        for _, (_, _, w, _, lbl) in self.SPREAD_ASSETS.items():
            weight_map[lbl] = w
        for _, (_, _, w, lbl) in self.MACRO_ASSETS.items():
            weight_map[lbl] = w

        records = []
        for col in self.signals.columns:
            val = row[col]
            w   = weight_map.get(col, float("nan"))
            if "Penalty" in col:
                status = "PENALTY ACTIVE" if val < 0 else "No Penalty"
            elif pd.isna(val):
                status = "N/A"
            else:
                status = "Risk On" if val > 0 else "Risk Off"
            records.append({"Signal": col, "Weight": w, "Value": round(val, 2),
                            "Status": status})

        df = pd.DataFrame(records).set_index("Signal")
        print(f"\nSignal snapshot — {dt}")
        print("─" * 50)
        print(df.to_string())
        total = row.sum()
        print(f"{'─' * 50}\nComposite Score : {total:.2f}")
        return df

    # ═════════════════════════════════════════════════════════════════════════
    # 6. Daily Snapshot  ← 오늘의 리스크 현황 (핵심 출력)
    # ═════════════════════════════════════════════════════════════════════════

    def daily_snapshot(self, date: str | None = None) -> dict:
        """
        오늘(또는 지정 날짜)의 리스크 현황을 콘솔에 출력하고 dict로 반환합니다.

        포함 항목
        ---------
        - 복합 스코어 & 레짐
        - 역사적 백분위 순위  (오늘이 전체 구간에서 몇 % 수준인가)
        - 30일 / 90일 평균 스코어 대비 변화
        - 개별 시그널 현황  (Risk On / Risk Off / Penalty 그룹 분류)

        Parameters
        ----------
        date : "YYYY-MM-DD" or None  → None이면 마지막 날짜(오늘)

        Returns
        -------
        dict  {date, score, regime, pct_rank, vs_30d, vs_90d, signals}
        """
        if self.scores is None or self.signals is None:
            raise RuntimeError("get_regime()를 먼저 실행하세요.")

        # ── 날짜 결정 ──────────────────────────────────────────────────────
        if date is None:
            dt = self.scores.index[-1]
        else:
            dt = pd.Timestamp(date)
            if dt not in self.scores.index:
                # 가장 가까운 이전 영업일로 대체
                dt = self.scores.index[self.scores.index.get_indexer(
                    [dt], method="ffill"
                )[0]]

        score  = float(self.scores.loc[dt])
        regime = str(self.regimes.loc[dt])

        # ── 역사적 맥락 ────────────────────────────────────────────────────
        pct_rank = float((self.scores <= score).mean() * 100)   # 백분위 순위

        pos = int(self.scores.index.get_loc(dt))
        avg_30d  = float(self.scores.iloc[max(0, pos - 30) : pos].mean()) \
                   if pos > 0 else float("nan")
        avg_90d  = float(self.scores.iloc[max(0, pos - 90) : pos].mean()) \
                   if pos > 0 else float("nan")
        vs_30d   = score - avg_30d  if not np.isnan(avg_30d) else float("nan")
        vs_90d   = score - avg_90d  if not np.isnan(avg_90d) else float("nan")

        # ── 시그널 분류 ────────────────────────────────────────────────────
        sig_row  = self.signals.loc[dt]

        on_sigs  = [(c, float(sig_row[c])) for c in sig_row.index
                    if float(sig_row[c]) > 0 and "Penalty" not in c]
        off_sigs = [(c, float(sig_row[c])) for c in sig_row.index
                    if float(sig_row[c]) == 0 and "Penalty" not in c]
        pen_sigs = [(c, float(sig_row[c])) for c in sig_row.index
                    if "Penalty" in c]

        max_score = 17.0   # 이론적 최대 (SKEW 패널티 없는 경우)
        bar_filled = max(int(score / max_score * 20), 0)
        score_bar  = "█" * bar_filled + "░" * (20 - bar_filled)

        # ── 레짐 색상 기호 ─────────────────────────────────────────────────
        REGIME_ICON = {
            "Risk On":    "🟢",
            "Transition": "🟡",
            "Risk Off":   "🔴",
        }
        icon = REGIME_ICON.get(regime, "⚪")

        # ── 콘솔 출력 ──────────────────────────────────────────────────────
        sep   = "═" * 64
        sep_s = "─" * 64
        print(f"\n{sep}")
        print(f"  오늘의 리스크 현황   {dt.strftime('%Y-%m-%d (%a)')}")
        print(sep)
        print(f"\n  {icon}  레짐 : [ {regime} ]")
        print(f"     스코어 : {score:.1f} / {max_score:.0f}   "
              f"[{score_bar}]")
        print(f"     역사적 백분위 : {pct_rank:.1f}%  "
              f"({'하위권' if pct_rank < 33 else '중간' if pct_rank < 67 else '상위권'})")

        def _delta(v):
            if np.isnan(v):
                return "N/A"
            sign = "▲" if v > 0 else "▼" if v < 0 else "━"
            return f"{sign} {abs(v):.2f}점"

        print(f"     30일 평균 대비  : {_delta(vs_30d)}   "
              f"(30d avg = {avg_30d:.1f})")
        print(f"     90일 평균 대비  : {_delta(vs_90d)}   "
              f"(90d avg = {avg_90d:.1f})")

        print(f"\n{sep_s}")
        print("  ✔  RISK ON  시그널")
        print(sep_s)
        if on_sigs:
            for name, val in on_sigs:
                print(f"     ✔  {name:<22}  +{val:.1f}")
        else:
            print("     (없음)")

        print(f"\n{sep_s}")
        print("  ✘  RISK OFF 시그널")
        print(sep_s)
        if off_sigs:
            for name, val in off_sigs:
                print(f"     ✘  {name:<22}   {val:.1f}")
        else:
            print("     (없음 — 전부 Risk On)")

        if pen_sigs:
            print(f"\n{sep_s}")
            print("  ⚠  PENALTY")
            print(sep_s)
            for name, val in pen_sigs:
                active = val < 0
                mark   = "⚠  발동 중!" if active else "   미발동"
                print(f"     {mark}  {name:<22}  {val:.1f}")

        print(f"\n{sep}")
        print(f"  복합 스코어 합계 : {score:.2f}   →   [ {regime} ]")
        print(f"{sep}\n")

        return {
            "date":     dt.strftime("%Y-%m-%d"),
            "score":    score,
            "regime":   regime,
            "pct_rank": pct_rank,
            "vs_30d":   vs_30d,
            "vs_90d":   vs_90d,
            "on_signals":  on_sigs,
            "off_signals": off_sigs,
            "penalties":   pen_sigs,
        }

    # ═════════════════════════════════════════════════════════════════════════
    # 7. Excel Export
    # ═════════════════════════════════════════════════════════════════════════

    def to_excel(self, path: str = "risk_score_report.xlsx") -> None:
        """
        보고용 Excel 파일을 생성합니다.

        Sheet 1 "Risk Score 결과"
            - 실행 요약  (기간 / 스코어 범위 / 최신 레짐)
            - 레짐 분포  (일수 & 비율, 막대 시각화)
            - 최신 시그널 현황  (자산별 당일 기여값 & 상태, 색상 코딩)
            - 전체 일별 스코어 이력  (레짐별 배경색, 자동 필터, 틀 고정)

        Sheet 2 "모델 정의"
            - 모델 개요 및 산출 로직 설명
            - 자산 유니버스 전체 스펙 테이블
            - 매크로 필터 테이블
            - SKEW 패널티 규칙
            - 레짐 분류 기준

        Parameters
        ----------
        path : str   저장 경로  (기본값: risk_score_report.xlsx)
        """
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(
                "openpyxl이 필요합니다:  pip install openpyxl"
            )

        if self.scores is None or self.regimes is None:
            raise RuntimeError("get_regime()를 먼저 실행하세요.")

        import openpyxl

        wb = openpyxl.Workbook()

        ws_today = wb.active
        ws_today.title = "오늘의 리스크 현황"

        ws_hist = wb.create_sheet("이력 비교")
        ws_def  = wb.create_sheet("모델 정의")

        self._excel_today_sheet(ws_today)
        self._excel_history_sheet(ws_hist)
        self._excel_definition_sheet(ws_def)

        wb.save(path)
        print(f"[INFO] Excel report saved → {path}")

    # ── Sheet 1: 오늘의 리스크 현황 ──────────────────────────────────────────

    def _excel_today_sheet(self, ws) -> None:
        """Sheet 1: 오늘의 리스크 현황 — 메인 보고 시트"""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        NAVY  = "1A3650";  DBLUE = "2E5077";  MBLUE = "3D7EAA"
        LBLUE = "EBF4FB";  LGREY = "F4F6F8";  WHITE = "FFFFFF"
        REG_BG = {"Risk On": "D5F5E3", "Transition": "FDEBD0",
                  "Risk Off": "FADBD8", "Unknown": "F5F5F5"}
        REG_FG = {"Risk On": "1D6A39", "Transition": "784212",
                  "Risk Off": "922B21", "Unknown": "555555"}
        ICON   = {"Risk On": "▲ Risk On", "Transition": "━ Transition",
                  "Risk Off": "▼ Risk Off"}

        FN = "맑은 고딕"
        def fnt(bold=False, sz=10, color="000000", italic=False):
            return Font(name=FN, bold=bold, size=sz, color=color, italic=italic)
        def bg(h):   return PatternFill("solid", fgColor=h)
        def al(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        _t = Side(style="thin", color="D0D0D0")
        _m = Side(style="medium", color="999999")
        def bdr():   return Border(left=_t, right=_t, top=_t, bottom=_t)
        def bdr_m(): return Border(left=_m, right=_m, top=_m, bottom=_m)

        def fill_row(r, n, h):
            for ci in range(1, n + 1): ws.cell(r, ci).fill = bg(h)

        def sec_hdr(row, title, n=6):
            fill_row(row, n, DBLUE)
            ws.merge_cells(f"A{row}:{chr(64+n)}{row}")
            c = ws.cell(row=row, column=1, value=title)
            c.font = fnt(True, 10, WHITE); c.fill = bg(DBLUE)
            c.alignment = al("left", "center")
            ws.row_dimensions[row].height = 20
            return row + 1

        def col_hdr(row, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = fnt(True, 9, WHITE); c.fill = bg(MBLUE)
                c.alignment = al("center", "center"); c.border = bdr()
            ws.row_dimensions[row].height = 18
            return row + 1

        # 열 너비
        for col, w in zip("ABCDEF", [24, 14, 16, 22, 14, 14]):
            ws.column_dimensions[col].width = w

        # ── 오늘 데이터 준비 ──────────────────────────────────────────────
        dt         = self.scores.index[-1]
        score      = float(self.scores.iloc[-1])
        regime     = str(self.regimes.iloc[-1])
        pct_rank   = float((self.scores <= score).mean() * 100)
        pos        = len(self.scores) - 1
        avg_30d    = float(self.scores.iloc[max(0, pos-30):pos].mean()) if pos > 0 else float("nan")
        avg_90d    = float(self.scores.iloc[max(0, pos-90):pos].mean()) if pos > 0 else float("nan")
        vs_30d     = score - avg_30d
        vs_90d     = score - avg_90d
        max_score  = 17.0
        bar_filled = max(int(score / max_score * 20), 0)
        score_bar  = "█" * bar_filled + "░" * (20 - bar_filled)

        sig_row = self.signals.iloc[-1]
        wmap: dict[str, float] = {}
        for _, (_, w, _, lbl) in self.PRICE_ASSETS.items(): wmap[lbl] = w
        wmap["SKEW Penalty"] = self.SKEW_PENALTY
        for _, (_, _, w, _, lbl) in self.SPREAD_ASSETS.items(): wmap[lbl] = w
        for _, (_, _, w, lbl) in self.MACRO_ASSETS.items(): wmap[lbl] = w

        row = 1

        # ── ① 제목 ──────────────────────────────────────────────────────
        for r in range(1, 4):
            fill_row(r, 6, NAVY)
        ws.row_dimensions[1].height = 26; ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 14
        ws.merge_cells("A1:F2")
        t = ws.cell(1, 1, "오늘의 리스크 현황")
        t.font = fnt(True, 18, WHITE); t.fill = bg(NAVY); t.alignment = al("center", "center")
        ws.merge_cells("A3:F3")
        s = ws.cell(3, 1, f"기준일: {dt.strftime('%Y-%m-%d (%A)')}    |    분석 기간: {self._start_date} ~ {self._end_date}    |    생성: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}")
        s.font = fnt(sz=9, color="BBBBBB", italic=True); s.fill = bg(NAVY); s.alignment = al("center", "center")
        row = 5

        # ── ② 레짐 배지 (가장 큰 강조) ───────────────────────────────────
        row = sec_hdr(row, "  ① 오늘의 레짐 & 스코어")
        rbg = REG_BG.get(regime, WHITE)
        rfg = REG_FG.get(regime, "000000")

        ws.merge_cells(f"A{row}:B{row+1}")
        label_c = ws.cell(row=row, column=1, value="레짐")
        label_c.font = fnt(True, 11, rfg); label_c.fill = bg(rbg)
        label_c.alignment = al("center", "center"); label_c.border = bdr_m()

        ws.merge_cells(f"C{row}:F{row+1}")
        regime_c = ws.cell(row=row, column=3, value=ICON.get(regime, regime))
        regime_c.font = fnt(True, 22, rfg); regime_c.fill = bg(rbg)
        regime_c.alignment = al("center", "center"); regime_c.border = bdr_m()
        ws.row_dimensions[row].height = 28; ws.row_dimensions[row+1].height = 28
        row += 2

        # 스코어 + 바
        for ci, (lbl, val, bold_col) in enumerate([
            ("스코어",    f"{score:.1f} / {max_score:.0f}",   True),
            ("스코어 바", score_bar,                           False),
            ("백분위",    f"{pct_rank:.1f}%  ({('하위권' if pct_rank < 33 else '중간' if pct_rank < 67 else '상위권')})", False),
        ], 0):
            ws.merge_cells(f"A{row}:B{row}")
            lc = ws.cell(row=row, column=1, value=lbl)
            lc.font = fnt(True, 9); lc.fill = bg(LBLUE)
            lc.alignment = al("left", "center"); lc.border = bdr()
            ws.merge_cells(f"C{row}:F{row}")
            vc = ws.cell(row=row, column=3, value=val)
            vc.font = fnt(bold_col, 10, rfg if bold_col else "222222")
            vc.fill = bg(rbg if bold_col else WHITE)
            vc.alignment = al("left" if ci > 0 else "center", "center")
            vc.border = bdr()
            ws.row_dimensions[row].height = 18; row += 1

        row += 1

        # ── ③ 역사적 맥락 ─────────────────────────────────────────────────
        row = sec_hdr(row, "  ② 역사적 맥락 비교  (오늘 스코어 vs 과거 평균)")
        row = col_hdr(row, ["구분", "스코어", "오늘 대비", "해석", "", ""])

        ctx_rows = [
            ("오늘",         score,   float("nan"),  f"기준"),
            ("30일 평균",    avg_30d, vs_30d,        "최근 1개월 흐름"),
            ("90일 평균",    avg_90d, vs_90d,        "최근 3개월 흐름"),
            ("역사적 최고",  float(self.scores.max()), score - float(self.scores.max()), "전체 구간 최고"),
            ("역사적 최저",  float(self.scores.min()), score - float(self.scores.min()), "전체 구간 최저"),
        ]
        for i, (lbl, val, diff, interp) in enumerate(ctx_rows):
            rbg_i = rbg if i == 0 else (WHITE if i % 2 == 0 else LGREY)
            def _diff_str(d):
                if np.isnan(d): return "─"
                return f"▲ +{d:.1f}" if d > 0 else f"▼ {d:.1f}" if d < 0 else "━  0.0"
            diff_color = ("1D6A39" if not np.isnan(diff) and diff > 0
                          else "922B21" if not np.isnan(diff) and diff < 0
                          else "555555")
            for ci, v in enumerate([lbl, f"{val:.1f}" if not np.isnan(val) else "─",
                                     _diff_str(diff), interp, "", ""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(rbg_i); c.border = bdr()
                c.alignment = al("center" if ci == 2 else "left", "center")
                c.font = (fnt(True, 9, rfg if i == 0 else diff_color)
                          if ci in (1, 3) else fnt(sz=9))
            ws.row_dimensions[row].height = 16; row += 1

        row += 1

        # ── ④ 시그널 현황 ─────────────────────────────────────────────────
        row = sec_hdr(row, f"  ③ 시그널별 현황  ({dt.strftime('%Y-%m-%d')})")
        row = col_hdr(row, ["시그널", "가중치", "당일 기여값", "상태", "조건 해석", ""])

        for i, sig_col in enumerate(self.signals.columns):
            val  = float(sig_row[sig_col])
            w    = wmap.get(sig_col, float("nan"))
            is_pen = "Penalty" in sig_col

            if is_pen:
                status = "⚠ PENALTY 발동" if val < 0 else "정상"
                sfg = REG_FG["Risk Off"] if val < 0 else REG_FG["Risk On"]
                sbg = REG_BG["Risk Off"] if val < 0 else WHITE
                interp = f"SKEW > {self.SKEW_THRESHOLD} → {self.SKEW_PENALTY}점 차감"
            elif val > 0:
                status = "✔ Risk On"; sfg = REG_FG["Risk On"]
                sbg = REG_BG["Risk On"] if i % 2 == 0 else "E8F8F0"
                interp = "MA 크로스 조건 충족"
            else:
                status = "✘ Risk Off"; sfg = REG_FG["Risk Off"]
                sbg = REG_BG["Risk Off"] if i % 2 == 0 else "FBEAEA"
                interp = "MA 크로스 조건 미충족"

            for ci, v in enumerate([sig_col, w, round(val, 2), status, interp, ""], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(sbg); c.border = bdr()
                c.alignment = al("center" if ci in (2, 3) else "left", "center")
                c.font = (fnt(True, 9, sfg) if ci == 4 else fnt(sz=9))
            ws.row_dimensions[row].height = 15; row += 1

        # 합계 행
        for ci, v in enumerate(
            ["합계 (Composite Score)", "", round(score, 2),
             ICON.get(regime, regime), "", ""], 1
        ):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = fnt(True, 11, rfg); c.fill = bg(rbg)
            c.border = bdr_m(); c.alignment = al("center" if ci > 1 else "left", "center")
        ws.row_dimensions[row].height = 22

        ws.freeze_panes = "A6"
        ws.sheet_properties.tabColor = REG_BG.get(regime, "2E5077")

    # ── Sheet 2: 이력 비교 ────────────────────────────────────────────────────

    def _excel_history_sheet(self, ws) -> None:
        """Sheet 2: 이력 비교 — 레짐 분포 요약 + 전체 일별 스코어 테이블"""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        NAVY = "1A3650"; DBLUE = "2E5077"; MBLUE = "3D7EAA"
        LBLUE = "EBF4FB"; WHITE = "FFFFFF"
        REG_BG = {"Risk On": "D5F5E3", "Transition": "FDEBD0",
                  "Risk Off": "FADBD8", "Unknown": "F5F5F5"}
        REG_FG = {"Risk On": "1D6A39", "Transition": "784212",
                  "Risk Off": "922B21", "Unknown": "555555"}

        FN = "맑은 고딕"
        def fnt(bold=False, sz=10, color="000000", italic=False):
            return Font(name=FN, bold=bold, size=sz, color=color, italic=italic)
        def bg(h):   return PatternFill("solid", fgColor=h)
        def al(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        _t = Side(style="thin", color="D0D0D0")
        def bdr(): return Border(left=_t, right=_t, top=_t, bottom=_t)

        def lighten(hex6, pct=0.45):
            r = int(int(hex6[0:2], 16) + (255 - int(hex6[0:2], 16)) * pct)
            g = int(int(hex6[2:4], 16) + (255 - int(hex6[2:4], 16)) * pct)
            b = int(int(hex6[4:6], 16) + (255 - int(hex6[4:6], 16)) * pct)
            return f"{r:02X}{g:02X}{b:02X}"

        def fill_row(r, n, h):
            for ci in range(1, n + 1): ws.cell(r, ci).fill = bg(h)

        def sec_hdr(row, title, n=5):
            fill_row(row, n, DBLUE)
            ws.merge_cells(f"A{row}:{chr(64+n)}{row}")
            c = ws.cell(row=row, column=1, value=title)
            c.font = fnt(True, 10, WHITE); c.fill = bg(DBLUE)
            c.alignment = al("left", "center")
            ws.row_dimensions[row].height = 20
            return row + 1

        def col_hdr(row, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = fnt(True, 9, WHITE); c.fill = bg(MBLUE)
                c.alignment = al("center", "center"); c.border = bdr()
            ws.row_dimensions[row].height = 18
            return row + 1

        for col, w in zip("ABCDE", [18, 11, 16, 34, 14]):
            ws.column_dimensions[col].width = w

        p_on  = float(self.scores.quantile(self.RISK_ON_PCT  / 100))
        p_off = float(self.scores.quantile(self.RISK_OFF_PCT / 100))

        row = 1

        # ── 제목 ──────────────────────────────────────────────────────────
        for r in range(1, 4): fill_row(r, 5, NAVY)
        ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 14
        ws.merge_cells("A1:E2")
        t = ws.cell(1, 1, "이력 비교  —  일별 리스크 스코어 전체 이력")
        t.font = fnt(True, 14, WHITE); t.fill = bg(NAVY); t.alignment = al("center", "center")
        ws.merge_cells("A3:E3")
        s = ws.cell(3, 1, f"분석 기간: {self._start_date} ~ {self._end_date}    |    "
                           f"Risk On(≥p{self.RISK_ON_PCT}): {p_on:.1f}    |    "
                           f"Risk Off(≤p{self.RISK_OFF_PCT}): {p_off:.1f}")
        s.font = fnt(sz=9, color="BBBBBB", italic=True); s.fill = bg(NAVY)
        s.alignment = al("center", "center")
        row = 5

        # ── 레짐 분포 요약 ────────────────────────────────────────────────
        row = sec_hdr(row, "  ① 레짐 분포 요약  (전체 분석 기간)")
        row = col_hdr(row, ["레짐", "일수", "비율", "막대", "최신 스코어 위치"])

        counts = self.regimes.value_counts(); total = int(counts.sum())
        latest_score = float(self.scores.iloc[-1])

        for regime in ["Risk On", "Transition", "Risk Off"]:
            cnt = int(counts.get(regime, 0)); pct = cnt / total * 100
            bar = "▇" * int(pct / 2.5)
            rfg = REG_FG.get(regime, "000000"); rbg = REG_BG.get(regime, WHITE)
            in_regime = "◀ 현재 위치" if str(self.regimes.iloc[-1]) == regime else ""
            for ci, v in enumerate([regime, cnt, f"{pct:.1f}%", bar, in_regime], 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = fnt(sz=9, bold=(ci == 1 or ci == 5), color=rfg)
                c.fill = bg(rbg); c.border = bdr()
                c.alignment = al("center" if ci > 1 else "left", "center")
            ws.row_dimensions[row].height = 16; row += 1

        # 통계 요약 행
        row += 1
        for label, val in [
            ("평균 스코어",    f"{self.scores.mean():.2f}"),
            ("표준편차",       f"{self.scores.std():.2f}"),
            ("최고",           f"{self.scores.max():.2f}"),
            ("최저",           f"{self.scores.min():.2f}"),
            ("오늘 스코어",    f"{latest_score:.2f}"),
            ("오늘 백분위",    f"{float((self.scores <= latest_score).mean()*100):.1f}%"),
        ]:
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = fnt(True, 9); lc.fill = bg(LBLUE)
            lc.alignment = al("left", "center"); lc.border = bdr()
            ws.merge_cells(f"B{row}:E{row}")
            vc = ws.cell(row=row, column=2, value=val)
            vc.font = fnt(sz=9); vc.alignment = al("left", "center"); vc.border = bdr()
            ws.row_dimensions[row].height = 15; row += 1

        row += 1

        # ── 전체 일별 스코어 이력 ─────────────────────────────────────────
        row = sec_hdr(row, "  ② 전체 일별 스코어 이력  (자동 필터 사용 가능 ▼)")
        tbl_start = row
        row = col_hdr(row, ["날짜", "스코어", "레짐", "스코어 바 (vs 최대값)", "오늘 대비"])

        max_sc = max(float(self.scores.max()), 1.0)

        for i, (dt, sc, rg) in enumerate(
            zip(self.scores.index, self.scores.values, self.regimes.values)
        ):
            sc_f   = float(sc)
            filled = max(int(sc_f / max_sc * 28), 0)
            bar    = "█" * filled + "░" * (28 - filled)
            diff   = sc_f - latest_score
            diff_s = (f"▲ +{diff:.1f}" if diff > 0
                      else f"▼ {diff:.1f}" if diff < 0 else "━  0.0")
            rbg    = REG_BG.get(rg, WHITE)
            row_bg = rbg if i % 2 == 0 else lighten(rbg, 0.5)
            is_today = (i == len(self.scores) - 1)

            for ci, v in enumerate(
                [dt.strftime("%Y-%m-%d"), round(sc_f, 2), rg, bar, diff_s], 1
            ):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = bg(row_bg); c.border = bdr()
                c.alignment = al("center" if ci != 1 else "left", "center")
                c.font = (fnt(True, 9, REG_FG.get(rg, "000000")) if ci == 3
                          else fnt(True, 9) if is_today
                          else fnt(sz=9))
            ws.row_dimensions[row].height = 14; row += 1

        ws.freeze_panes = f"A{tbl_start + 1}"
        ws.auto_filter.ref = f"A{tbl_start}:E{row - 1}"
        ws.sheet_properties.tabColor = "2E5077"

    # ── Sheet 3 helper ────────────────────────────────────────────────────────

    def _excel_definition_sheet(self, ws) -> None:
        """Sheet 2: 모델 정의 및 산출 방식 시트 작성"""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        # ── 색상 ───────────────────────────────────────────────────────────
        NAVY   = "1A3650"
        DBLUE  = "2E5077"
        MBLUE  = "3D7EAA"
        LBLUE  = "EBF4FB"
        LGREY  = "F4F6F8"
        WHITE  = "FFFFFF"
        GREEN  = "D5F5E3"
        ORANGE = "FDEBD0"
        RED    = "FADBD8"
        FG_G   = "1D6A39"
        FG_O   = "784212"
        FG_R   = "922B21"

        FN = "맑은 고딕"

        def fnt(bold=False, sz=10, color="000000", italic=False):
            return Font(name=FN, bold=bold, size=sz, color=color, italic=italic)

        def bg(h):
            return PatternFill("solid", fgColor=h)

        def al(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        _thin = Side(style="thin", color="D0D0D0")

        def bdr():
            return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        # ── 열 너비 ────────────────────────────────────────────────────────
        col_w = {"A": 14, "B": 18, "C": 22, "D": 18,
                 "E": 10,  "F": 10,  "G": 26, "H": 30}
        for col, w in col_w.items():
            ws.column_dimensions[col].width = w

        def fill_row(ws, r, n, hex6):
            for ci in range(1, n + 1):
                ws.cell(r, ci).fill = bg(hex6)

        def section_header(ws, row, title, n_cols=8):
            fill_row(ws, row, n_cols, DBLUE)
            ws.merge_cells(f"A{row}:{chr(64 + n_cols)}{row}")
            c = ws.cell(row=row, column=1, value=title)
            c.font      = fnt(bold=True, sz=10, color=WHITE)
            c.fill      = bg(DBLUE)
            c.alignment = al("left", "center")
            ws.row_dimensions[row].height = 20
            return row + 1

        def col_headers(ws, row, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font      = fnt(bold=True, sz=9, color=WHITE)
                c.fill      = bg(MBLUE)
                c.alignment = al("center", "center")
                c.border    = bdr()
            ws.row_dimensions[row].height = 18
            return row + 1

        def text_row(ws, row, text, indent=0, bold=False,
                     col_span=8, row_bg=WHITE, text_color="000000"):
            ws.merge_cells(f"A{row}:{chr(64 + col_span)}{row}")
            c = ws.cell(row=row, column=1,
                        value=("    " * indent) + text)
            c.font      = fnt(bold=bold, sz=9, color=text_color)
            c.fill      = bg(row_bg)
            c.alignment = al("left", "center", wrap=True)
            ws.row_dimensions[row].height = 16
            return row + 1

        row = 1

        # ──────────────────────────────────────────────────────────────────
        # 제목
        # ──────────────────────────────────────────────────────────────────
        for r in range(1, 4):
            fill_row(ws, r, 8, NAVY)
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 14

        ws.merge_cells("A1:H2")
        t = ws.cell(row=1, column=1,
                    value="모델 정의 및 산출 방식")
        t.font      = fnt(bold=True, sz=16, color=WHITE)
        t.fill      = bg(NAVY)
        t.alignment = al("center", "center")

        ws.merge_cells("A3:H3")
        s = ws.cell(row=3, column=1,
                    value="Global Multi-Asset Risk Scoring Model  —  Logic Reference")
        s.font      = fnt(sz=9, color="BBBBBB", italic=True)
        s.fill      = bg(NAVY)
        s.alignment = al("center", "center")
        row = 5

        # ──────────────────────────────────────────────────────────────────
        # ① 모델 개요
        # ──────────────────────────────────────────────────────────────────
        row = section_header(ws, row, "  ① 모델 개요")

        overview_lines = [
            ("본 모델은 글로벌 자산시장의 Risk On / Risk Off 레짐을 매일 수치화하는 "
             "복합 스코어링 시스템입니다.", False, WHITE),
            ("FX, 신흥국 채권/통화, 금리, 주식, 원자재, 변동성, 신용 등 12개 "
             "캐너리 자산과 3개 거시 지표를 블룸버그 터미널에서 직접 수집합니다.", False, WHITE),
            ("각 자산의 이동평균(MA) 크로스오버 또는 임계값 조건으로 "
             "Risk On 시그널(0 또는 +가중치)을 생성하고, "
             "SKEW Index는 패널티(−2점)로 반영합니다.", False, WHITE),
            ("복합 스코어 = Σ(개별 가중 시그널) + SKEW 패널티", True, DBLUE),
            ("이론적 최대: +17점  |  이론적 최솟값: −2점  (SKEW 패널티 발동 시)", False, LGREY),
        ]
        for text, bold, row_bg in overview_lines:
            row = text_row(ws, row, text, bold=bold, row_bg=row_bg,
                           text_color=WHITE if row_bg == DBLUE else "222222")

        row += 1

        # ──────────────────────────────────────────────────────────────────
        # ② 자산 유니버스
        # ──────────────────────────────────────────────────────────────────
        row = section_header(ws, row, "  ② 캐너리 자산 유니버스 (가격 기반 시그널)")

        hdrs = ["분류", "자산명 (한글)", "블룸버그 티커",
                "블룸버그 필드", "MA 기간", "가중치",
                "Risk On 조건", "비고"]
        row = col_headers(ws, row, hdrs)

        # 자산 정보 매핑
        CAT_MAP = {
            "DXY Curncy":     ("FX",      "달러 인덱스"),
            "EMB US Equity":  ("EM",      "EM 채권"),
            "CEW US Equity":  ("EM",      "EM 통화"),
            "BND US Equity":  ("금리",    "미국채 종합"),
            "TIP US Equity":  ("금리",    "TIPS"),
            "VEA US Equity":  ("주식",    "선진국 주식"),
            "GLD US Equity":  ("원자재",  "금"),
            "DBC US Equity":  ("원자재",  "원자재"),
            "VIX Index":      ("변동성",  "VIX"),
            "VVIX Index":     ("변동성",  "VVIX"),
        }
        NOTE_MAP = {
            "DXY Curncy":    "달러 약세 = Risk On / 가중치 2배 (고감도)",
            "GLD US Equity": "안전자산 자금 유출 = Risk On",
            "BND US Equity": "금리 상승(채권 하락) = Risk On 맥락",
            "VIX Index":     "가중치 2배 (고감도)",
        }

        asset_rows_data = []
        for ticker, (ma_p, w, direction, label) in self.PRICE_ASSETS.items():
            cat, kor = CAT_MAP.get(ticker, ("기타", label))
            cond = f"가격 {'<' if direction == 'below' else '>'} {ma_p}일 MA"
            asset_rows_data.append(
                [cat, kor, ticker, "PX_LAST", f"{ma_p}일", w, cond,
                 NOTE_MAP.get(ticker, "")]
            )

        # SKEW (별도 처리)
        asset_rows_data.append(
            ["꼬리리스크", "SKEW Index", self.SKEW_TICKER, "PX_LAST",
             "—  (MA 미사용)", f"{self.SKEW_PENALTY}  (패널티)",
             f"SKEW > {self.SKEW_THRESHOLD}  →  −2점 차감",
             "임계값 기반 / 다른 시그널과 독립적으로 작동"]
        )

        # Spread assets
        for ticker, (field, ma_p, w, direction, label) in self.SPREAD_ASSETS.items():
            cond = f"스프레드 {'<' if direction == 'below' else '>'} {ma_p}일 MA"
            asset_rows_data.append(
                ["신용", "미국 HY OAS", ticker, field,
                 f"{ma_p}일", w, cond,
                 "스프레드 축소(위험 선호) = Risk On / 가중치 2배"]
            )

        row_bgs = [LGREY, WHITE]
        for i, r_data in enumerate(asset_rows_data):
            rbg = row_bgs[i % 2]
            # SKEW 행은 연한 빨강으로 구분
            if "패널티" in str(r_data[5]):
                rbg = RED
            for ci, v in enumerate(r_data, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill      = bg(rbg)
                c.border    = bdr()
                c.alignment = al("center" if ci in (5, 6) else "left",
                                 "center", wrap=(ci == 8))
                c.font      = fnt(sz=9, bold=(ci == 6),
                                  color=(FG_R if "패널티" in str(r_data[5]) else "222222"))
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1

        # ──────────────────────────────────────────────────────────────────
        # ③ 매크로 필터
        # ──────────────────────────────────────────────────────────────────
        row = section_header(ws, row, "  ③ 매크로 필터  (월별 발표 → ffill 로 일별 적용)")

        macro_hdrs = ["지표명 (한글)", "블룸버그 티커", "블룸버그 필드",
                      "Risk On 조건", "가중치", "처리 방식", "", ""]
        row = col_headers(ws, row, macro_hdrs)

        MACRO_KOR = {
            "NAPMPMI Index": ("ISM 제조업 PMI",  "값 > 50"),
            "CPI YOY Index": ("CPI YoY (%)",    "값 < 3.0%"),
            "USURTOT Index": ("미국 실업률 (%)",  "값 < 4.0%"),
        }

        for i, (ticker, (cond, thr, w, label)) in \
                enumerate(self.MACRO_ASSETS.items()):
            kor_name, cond_str = MACRO_KOR.get(ticker, (label, f"{cond} {thr}"))
            rbg = WHITE if i % 2 == 0 else LGREY
            row_d = [kor_name, ticker, "PX_LAST", cond_str, w,
                     "월별 발표 → ffill → 일별 매핑", "", ""]
            for ci, v in enumerate(row_d, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill      = bg(rbg)
                c.border    = bdr()
                c.alignment = al("center" if ci == 5 else "left", "center")
                c.font      = fnt(sz=9, bold=(ci == 5))
            ws.row_dimensions[row].height = 16
            row += 1

        row = text_row(ws, row,
                       "※ Bloomberg BDH는 발표일에만 값을 반환(sparse). "
                       "코드에서 일별 인덱스로 reindex 후 ffill()로 "
                       "다음 발표일까지 값을 유지합니다.",
                       row_bg=ORANGE, text_color=FG_O)
        row += 1

        # ──────────────────────────────────────────────────────────────────
        # ④ SKEW 패널티 규칙
        # ──────────────────────────────────────────────────────────────────
        row = section_header(ws, row, "  ④ SKEW Index 패널티 로직")

        skew_lines = [
            ("SKEW Index는 S&P 500 옵션 시장의 꼬리 리스크(Tail Risk) 수요를 나타내는 지수입니다.", False, WHITE),
            ("다른 시그널과 달리 이동평균 비교가 아닌 절대 임계값 기반으로 작동합니다.", False, WHITE),
            (f"규칙:  SKEW > {self.SKEW_THRESHOLD}  →  스코어에서 {abs(self.SKEW_PENALTY):.0f}점 차감  (SKEW_PENALTY = {self.SKEW_PENALTY})", True, RED),
            (f"규칙:  SKEW ≤ {self.SKEW_THRESHOLD}  →  기여값 = 0  (패널티 없음)", True, GREEN),
            ("효과: 다른 지표가 모두 Risk On을 가리켜도 꼬리 리스크가 높으면 "
             "복합 스코어를 강제로 하향시킵니다.", False, LGREY),
        ]
        for text, bold, row_bg in skew_lines:
            color = (FG_R if row_bg == RED
                     else FG_G if row_bg == GREEN
                     else "222222")
            row = text_row(ws, row, text, bold=bold,
                           row_bg=row_bg, text_color=color)
        row += 1

        # ──────────────────────────────────────────────────────────────────
        # ⑤ 레짐 분류 기준
        # ──────────────────────────────────────────────────────────────────
        row = section_header(ws, row, "  ⑤ 레짐 분류 기준")
        row = col_headers(ws, row, ["레짐", "조건", "해석",
                                    "배경색 (차트)", "", "", "", ""])

        regime_def = [
            ("Risk On",    f"스코어 ≥ p{self.RISK_ON_PCT}  (역사적 상위 {100-self.RISK_ON_PCT}%)",
             "위험자산 선호 환경 — 주식·EM·원자재 강세 맥락",
             "초록색", GREEN, FG_G),
            ("Transition", f"p{self.RISK_OFF_PCT} < 스코어 < p{self.RISK_ON_PCT}",
             "중립·관망 구간 — 방향성 불명확",
             "주황색", ORANGE, FG_O),
            ("Risk Off",   f"스코어 ≤ p{self.RISK_OFF_PCT}  (역사적 하위 {self.RISK_OFF_PCT}%)",
             "안전자산 선호 환경 — 달러·국채·금 강세 맥락",
             "빨간색", RED, FG_R),
        ]
        for regime, cond, interp, color_name, rbg, rfg in regime_def:
            for ci, v in enumerate(
                [regime, cond, interp, color_name, "", "", "", ""], 1
            ):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill      = bg(rbg)
                c.border    = bdr()
                c.alignment = al("left", "center", wrap=(ci == 3))
                c.font      = fnt(sz=9, bold=(ci == 1), color=rfg)
            ws.row_dimensions[row].height = 20
            row += 1

        row = text_row(
            ws, row,
            f"※ 임계값은 고정값이 아닌 분석 기간 전체의 백분위로 자동 계산됩니다. "
            f"(현재 p{self.RISK_ON_PCT} / p{self.RISK_OFF_PCT} 기준)",
            row_bg=LBLUE, text_color="2E5077"
        )

        ws.sheet_properties.tabColor = "1A3650"


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # ── 날짜별 출력 폴더 생성 ─────────────────────────────────────────
    today_str = datetime.now().strftime("%Y-%m-%d")
    out_dir = os.path.join("canaria_risk_score_output", today_str)
    os.makedirs(out_dir, exist_ok=True)

    model = RiskScoringModel()

    # end_date 기본값 = 오늘, start_date 기본값 = 2년 전
    # 필요시 직접 지정:  model.run(start_date="2020-01-01", end_date="2025-02-19")
    results = model.run(plot=False)

    # ── 핵심 출력: 오늘의 리스크 현황 ──────────────────────────────────
    model.daily_snapshot()

    # ── 보조 출력: 최근 10일 이력 ──────────────────────────────────────
    print("최근 10 영업일:")
    print(results.tail(10).to_string())

    # ── 차트 저장 ────────────────────────────────────────────────────────
    chart_path = os.path.join(out_dir, "risk_score_chart.png")
    model.plot_results(save_path=chart_path)

    # ── Excel 보고서 저장 ───────────────────────────────────────────────
    today_str = model.scores.index[-1].strftime("%Y%m%d")
    excel_path = os.path.join(out_dir, f"risk_score_report_{today_str}.xlsx")
    model.to_excel(excel_path)
